""" 
 Maxim Dumortier
 Feb. 2017
 Cerisic, Mons, Belgium
 
 This class is used to export the data in the wanted way
 Available : 
  - CSV
  - Xlsx ? TODO
  - odt ? TODO
  - XML ? TODO

"""

from enum import Enum
import csv
from openpyxl import Workbook, load_workbook # pip install openpyxls


class Type(Enum):
    """
    Enum of the diffrent possible types of exporting
    """
    
    CSV     = 0
    XLSX    = 1
    XML     = 2


class Export:

    def __init__(self, file_path, type = "csv", header = "" ):
        # TODO : consider the case that the file path is given with the extention e.g. : myfile.csv
        # In this case, juste catch the extention and set it as type (ignore the setted type)
                
        self.file_path = file_path + "." + type.lower()
        self.firstwrittedline = True
                
        if type.lower() == 'csv':
            self.type = Type.CSV
            source = open(self.file_path,'w') # Just create the file and close it
            source.close()
                        
        elif type.lower() == 'xlsx':
            self.type = Type.XLSX 
            wb = Workbook()# Just create the file and close it
            ws = wb.active
            ws.title = "(PS)²DAq Data"
            wb.save(self.file_path)
            
        elif type.lower() == 'xml':
            self.type = Type.XML
            raise Exception("Not supported yet")
            
        else:
            raise ValueError('Unknown type : ', type)
            
        
        if header != "":
            self.writerow(header) # Just write the header as the first line in the file


    def writerow(self, line):
        if self.type == Type.CSV:
            with open(self.file_path,'a') as source:
                writer = csv.writer(source, delimiter=';', lineterminator='\n')
                writer.writerow(line)
                
        elif self.type == Type.XLSX:
            wb = load_workbook(self.file_path)
            ws = wb.active
            if self.firstwrittedline: # Use this little stuff in order to begin the append at the first line and not the second ...
                ws._current_row = 0
                self.firstwrittedline = False
            ws.append(line)
            wb.save(self.file_path)
            
            
        
